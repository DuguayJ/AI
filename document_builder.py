#!/usr/bin/env python3
"""
Universal Document Builder
Generates professional Word (.docx) or Excel (.xlsx) documents
from structured templates with consistent branding.

Usage:
    python document_builder.py                  # interactive wizard
    python document_builder.py --json data.json # from JSON file
    python document_builder.py --example        # write example JSON and exit
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import date as dt_date
from pathlib import Path
from typing import Any

# ── Excel dependencies ──────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── Brand palette ──────────────────────────────────────────────────────────
C = dict(
    dark_blue   ="1F3864",
    mid_blue    ="2E5FA3",
    light_blue  ="D6E4F7",
    accent_blue ="4A90D9",
    dark_gray   ="2C2C2C",
    med_gray    ="595959",
    light_gray  ="F4F6F9",
    border_gray ="D0D7E2",
    white       ="FFFFFF",
    warn_yellow ="FEF9EC",
    error_pink  ="FDF0F0",
    success_green="F0FAF4",
    teal        ="1A7F8E",
)

STATUS_COLORS = {
    "Completed":      C["teal"],
    "In Progress":    "8A6000",
    "Pending Review": C["mid_blue"],
    "Draft":          C["med_gray"],
}

# ─── Excel builder ──────────────────────────────────────────────────────────

def _thin_border(color: str = "D0D7E2") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(size: int = 11, bold: bool = True, color: str = "FFFFFF") -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)

def _body_font(size: int = 10, bold: bool = False, color: str = "2C2C2C") -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _center(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def build_excel(data: dict, out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    row = 1

    # ── Title bar ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = data.get("reportTitle", "Document Title")
    c.font = Font(name="Arial", size=18, bold=True, color=C["white"])
    c.fill = _fill(C["dark_blue"])
    c.alignment = _center()
    ws.row_dimensions[row].height = 40
    row += 1

    # ── Subtitle / doc type ────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = data.get("subtitle") or data.get("documentType", "")
    c.font = Font(name="Arial", size=12, color=C["mid_blue"])
    c.fill = _fill(C["light_blue"])
    c.alignment = _center()
    ws.row_dimensions[row].height = 24
    row += 1

    # ── Metadata strip ─────────────────────────────────────────────────────
    meta_pairs = [
        ("Organization", data.get("organization", "—")),
        ("Environment",  data.get("environment", "—")),
        ("Date",         data.get("date", dt_date.today().isoformat())),
        ("Prepared By",  data.get("preparedBy", "—")),
        ("Classification", data.get("classification", "Confidential")),
        ("Status",       data.get("status", "—")),
    ]
    for i in range(0, len(meta_pairs), 3):
        chunk = meta_pairs[i:i+3]
        for j, (label, val) in enumerate(chunk):
            col_l = j * 2 + 1
            col_v = col_l + 1
            lc = ws.cell(row=row, column=col_l, value=label)
            lc.font = _header_font(size=9, color=C["dark_blue"])
            lc.fill = _fill(C["light_gray"])
            lc.alignment = _left(wrap=False)
            lc.border = _thin_border()

            vc = ws.cell(row=row, column=col_v, value=val)
            vc.font = _body_font(size=9)
            vc.fill = _fill(C["white"])
            vc.alignment = _left(wrap=False)
            vc.border = _thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # blank spacer

    # ── Sections ───────────────────────────────────────────────────────────
    for sec in data.get("sections", []):
        # Section heading
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = sec.get("title", "Section")
        c.font = Font(name="Arial", size=13, bold=True, color=C["white"])
        c.fill = _fill(C["mid_blue"])
        c.alignment = _left(wrap=False)
        c.border = Border(bottom=Side(style="medium", color=C["accent_blue"]))
        ws.row_dimensions[row].height = 26
        row += 1

        for block in sec.get("blocks", []):
            btype = block.get("type")

            if btype == "text":
                ws.merge_cells(f"A{row}:F{row}")
                c = ws[f"A{row}"]
                c.value = block.get("content", "")
                c.font = _body_font(size=10)
                c.alignment = _left(wrap=True)
                ws.row_dimensions[row].height = 30
                row += 1

            elif btype in ("bullets", "numbered"):
                for idx, item in enumerate(block.get("items", []), start=1):
                    prefix = f"{idx}." if btype == "numbered" else "•"
                    ws.merge_cells(f"A{row}:F{row}")
                    c = ws[f"A{row}"]
                    c.value = f"    {prefix}  {item}"
                    c.font = _body_font(size=10)
                    c.fill = _fill(C["white"] if idx % 2 == 0 else C["light_gray"])
                    c.alignment = _left(wrap=True)
                    ws.row_dimensions[row].height = 22
                    row += 1

            elif btype == "callout":
                style_map = {
                    "info":    (C["light_blue"],    C["mid_blue"],  "ℹ INFO"),
                    "warning": (C["warn_yellow"],   "8A6000",       "⚠ WARNING"),
                    "error":   (C["error_pink"],    "B03030",       "✕ CRITICAL"),
                    "success": (C["success_green"], C["teal"],      "✓ NOTE"),
                }
                fill_c, txt_c, lbl = style_map.get(block.get("style", "info"), style_map["info"])
                ws.merge_cells(f"A{row}:F{row}")
                c = ws[f"A{row}"]
                c.value = f"{lbl}: {block.get('content', '')}"
                c.font = Font(name="Arial", size=10, bold=True, color=txt_c)
                c.fill = _fill(fill_c)
                c.alignment = _left(wrap=True)
                ws.row_dimensions[row].height = 30
                row += 1

            elif btype == "table":
                cols = block.get("columns", [])
                rows_data = block.get("rows", [])
                for ci, col in enumerate(cols, start=1):
                    c = ws.cell(row=row, column=ci, value=col)
                    c.font = _header_font(size=10, color=C["white"])
                    c.fill = _fill(C["dark_blue"])
                    c.alignment = _center(wrap=True)
                    c.border = _thin_border(C["mid_blue"])
                ws.row_dimensions[row].height = 22
                row += 1
                for ri, data_row in enumerate(rows_data):
                    for ci, val in enumerate(data_row, start=1):
                        c = ws.cell(row=row, column=ci, value=str(val))
                        c.font = _body_font(size=10)
                        c.fill = _fill(C["white"] if ri % 2 == 0 else C["light_gray"])
                        c.alignment = _left(wrap=True)
                        c.border = _thin_border()
                    ws.row_dimensions[row].height = 22
                    row += 1

            elif btype == "key_value":
                for ki, (k, v) in enumerate(block.get("items", [])):
                    lc = ws.cell(row=row, column=1, value=k)
                    lc.font = _body_font(size=10, bold=True, color=C["dark_blue"])
                    lc.fill = _fill(C["light_gray"] if ki % 2 == 0 else C["white"])
                    lc.alignment = _left(wrap=False)
                    lc.border = _thin_border()

                    ws.merge_cells(f"B{row}:F{row}")
                    vc = ws.cell(row=row, column=2, value=v)
                    vc.font = _body_font(size=10)
                    vc.fill = _fill(C["white"] if ki % 2 == 0 else C["light_gray"])
                    vc.alignment = _left(wrap=True)
                    vc.border = _thin_border()
                    ws.row_dimensions[row].height = 22
                    row += 1

            elif btype == "heading2":
                ws.merge_cells(f"A{row}:F{row}")
                c = ws[f"A{row}"]
                c.value = block.get("content", "")
                c.font = Font(name="Arial", size=11, bold=True, color=C["mid_blue"])
                c.fill = _fill(C["light_blue"])
                c.alignment = _left(wrap=False)
                ws.row_dimensions[row].height = 22
                row += 1

        row += 1  # section spacer

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = [30, 20, 20, 20, 20, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ─── Word builder (delegates to Node.js) ────────────────────────────────────

def build_word(data: dict, out_path: Path) -> Path:
    script = Path(__file__).parent / "build_word.js"
    if not script.exists():
        raise FileNotFoundError(f"build_word.js not found at {script}")
    result = subprocess.run(
        ["node", str(script), json.dumps(data), str(out_path)],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"build_word.js failed:\n{result.stderr}")
    return out_path


# ─── Interactive wizard ──────────────────────────────────────────────────────

def _ask(prompt: str, default: str = "") -> str:
    hint = f" [{default}]" if default else ""
    v = input(f"  {prompt}{hint}: ").strip()
    return v or default

def _ask_int(prompt: str, default: int = 0) -> int:
    while True:
        v = _ask(prompt, str(default))
        try:
            return int(v)
        except ValueError:
            print("  ↳ Please enter a number.")

def _banner(text: str) -> None:
    w = 60
    print(f"\n{'─' * w}")
    print(f"  {text}")
    print(f"{'─' * w}")


def _collect_block() -> dict | None:
    _banner("Add a content block  (leave type blank to finish section)")
    TYPES = ["text", "bullets", "numbered", "table", "key_value", "callout", "heading2"]
    print("  Block types: " + "  ·  ".join(TYPES))
    btype = _ask("Type", "")
    if not btype:
        return None

    block: dict[str, Any] = {"type": btype}

    if btype in ("text", "heading2"):
        block["content"] = _ask("Content")

    elif btype in ("bullets", "numbered"):
        items = []
        print("  Enter items, blank line to stop:")
        while True:
            item = _ask(f"  Item {len(items)+1}", "")
            if not item:
                break
            items.append(item)
        block["items"] = items

    elif btype == "callout":
        print("  Styles: info  warning  error  success")
        block["style"] = _ask("Style", "info")
        block["content"] = _ask("Content")

    elif btype == "table":
        cols_raw = _ask("Columns (comma-separated)")
        block["columns"] = [c.strip() for c in cols_raw.split(",") if c.strip()]
        n_rows = _ask_int(f"Number of data rows", 2)
        rows = []
        for r in range(n_rows):
            row_vals = []
            for col in block["columns"]:
                row_vals.append(_ask(f"  Row {r+1} · {col}"))
            rows.append(row_vals)
        block["rows"] = rows

    elif btype == "key_value":
        items = []
        print("  Enter key/value pairs, blank key to stop:")
        while True:
            k = _ask(f"  Key {len(items)+1}", "")
            if not k:
                break
            v = _ask(f"  Value for '{k}'")
            items.append([k, v])
        block["items"] = items

    return block


def _collect_section() -> dict:
    title = _ask("Section title")
    blocks = []
    while True:
        b = _collect_block()
        if b is None:
            break
        blocks.append(b)
    return {"title": title, "blocks": blocks}


def interactive_wizard() -> tuple[str, dict, Path]:
    _banner("Universal Document Builder  —  Interactive Wizard")

    fmt = _ask("Output format (word / excel)", "word").lower()
    while fmt not in ("word", "excel"):
        fmt = _ask("Please enter word or excel", "word").lower()

    data: dict[str, Any] = {}

    _banner("Document metadata")
    data["reportTitle"]   = _ask("Report / Document title", "Document Title")
    data["documentType"]  = _ask("Document type (e.g. IT Security, Asset Registry, Procedure)", "")
    data["subtitle"]      = _ask("Subtitle / scope", "")
    data["organization"]  = _ask("Organization / company", "")
    data["environment"]   = _ask("Environment / site", "")
    data["date"]          = _ask("Date", dt_date.today().isoformat())
    data["classification"]= _ask("Classification", "Confidential — Internal Use Only")
    data["preparedBy"]    = _ask("Prepared by", "")
    data["version"]       = _ask("Version", "1.0")
    print("  Statuses: Draft  In Progress  Pending Review  Completed")
    data["status"]        = _ask("Status", "Draft")

    if fmt == "word":
        data["includeToc"] = _ask("Include Table of Contents? (y/n)", "n").lower() == "y"

    _banner("Document sections")
    n_sections = _ask_int("How many sections?", 3)
    sections = []
    for i in range(n_sections):
        _banner(f"Section {i+1} of {n_sections}")
        sections.append(_collect_section())
    data["sections"] = sections

    safe_title = data["reportTitle"].lower().replace(" ", "_")[:40]
    default_out = f"output/{safe_title}.{fmt if fmt == 'xlsx' else 'docx'}".replace("word", "docx")
    out_raw = _ask("Output path", f"output/{safe_title}.{'xlsx' if fmt == 'excel' else 'docx'}")
    out_path = Path(out_raw)

    return fmt, data, out_path


# ─── Example JSON ────────────────────────────────────────────────────────────

EXAMPLE_DATA = {
    "reportTitle": "Asset Management Register",
    "documentType": "Asset Registry",
    "subtitle": "Hardware & Software Inventory — FY 2025",
    "organization": "Acme Corp",
    "environment": "All Sites",
    "date": dt_date.today().isoformat(),
    "classification": "Internal Use Only",
    "preparedBy": "IT Operations Team",
    "version": "1.0",
    "status": "In Progress",
    "includeToc": False,
    "sections": [
        {
            "title": "Overview",
            "blocks": [
                {"type": "text", "content": "This register tracks all IT hardware and software assets across Acme Corp sites."},
                {"type": "key_value", "items": [
                    ["Total Assets", "1,247"],
                    ["Sites Covered", "3 (NDH, HPH, HQ)"],
                    ["Last Audit", "April 15, 2025"],
                    ["Next Audit Due", "October 15, 2025"]
                ]},
                {"type": "callout", "style": "warning", "content": "142 assets are approaching end-of-life within 90 days. Refresh planning required."}
            ]
        },
        {
            "title": "Hardware Inventory",
            "blocks": [
                {"type": "table",
                 "columns": ["Asset Type", "Count", "Site", "Lifecycle Status"],
                 "rows": [
                     ["Workstations", "320", "All", "Active"],
                     ["Servers (Physical)", "48", "NDH / HPH", "Active"],
                     ["Network Switches", "67", "All", "Mixed"],
                     ["UPS Units", "22", "NDH / HPH", "Aging"]
                 ]}
            ]
        },
        {
            "title": "Next Steps",
            "blocks": [
                {"type": "numbered", "items": [
                    "Initiate refresh procurement for end-of-life workstations — Owner: Ops, Due: June 1",
                    "Schedule decommission of legacy servers — Owner: Infra, Due: July 15",
                    "Update CMDB with audit results — Owner: ITSM, Due: May 10"
                ]}
            ]
        }
    ]
}


# ─── Entry point ─────────────────────────────────────────────────────────────

def main() -> None:
    if "--example" in sys.argv:
        ex_path = Path("example_input.json")
        ex_path.write_text(json.dumps(EXAMPLE_DATA, indent=2))
        print(f"Example JSON written to: {ex_path}")
        print("Run:  python document_builder.py --json example_input.json")
        return

    if "--json" in sys.argv:
        idx = sys.argv.index("--json")
        json_path = Path(sys.argv[idx + 1])
        data = json.loads(json_path.read_text())
        fmt = _ask("Output format (word / excel)", "word").lower()
        safe = data.get("reportTitle", "document").lower().replace(" ", "_")[:40]
        ext = "xlsx" if fmt == "excel" else "docx"
        out_path = Path(_ask("Output path", f"output/{safe}.{ext}"))
    else:
        fmt, data, out_path = interactive_wizard()

    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\n  Building {fmt.upper()} document…")
    if fmt == "excel":
        result = build_excel(data, out_path)
    else:
        result = build_word(data, out_path)

    print(f"\n  ✓ Document generated: {result}\n")


if __name__ == "__main__":
    main()
